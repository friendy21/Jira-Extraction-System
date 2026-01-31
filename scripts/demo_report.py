#!/usr/bin/env python
"""
Demo Report Generator
Creates a sample report without requiring database connection.
"""

import sys
from pathlib import Path
from datetime import datetime, timedelta
from decimal import Decimal

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

from src.reports.charts import ChartBuilder
from src.utils.logger import setup_logging, get_logger

setup_logging()
logger = get_logger(__name__)


def create_demo_report():
    """Create a demo Excel report with mock data."""
    
    output_path = Path('./outputs')
    output_path.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = output_path / f"demo_dashboard_{timestamp}.xlsx"
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create Executive Summary
    ws = wb.create_sheet("Executive Summary")
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Jira Dashboard - Demo Report"
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=9)
    
    # Key Metrics
    row = 4
    ws[f'A{row}'] = "Key Metrics (Last 30 Days)"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E79")
    row += 2
    
    metrics = [
        ("Tickets Created", 156),
        ("Tickets Resolved", 142),
        ("Resolution Rate", "91.0%"),
        ("Average Cycle Time", "3.2 days"),
        ("Team Velocity", "45 points"),
    ]
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for metric_name, metric_value in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = metric_value
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        row += 1
    
    row += 2
    
    # Sprint Velocity
    ws[f'A{row}'] = "Sprint Velocity (Last 5 Sprints)"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E79")
    row += 2
    
    # Headers
    headers = ['Sprint', 'Committed', 'Completed', 'Velocity']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Mock sprint data
    data_start_row = row
    sprints = [
        ("Sprint 20", 42, 38, 38),
        ("Sprint 21", 45, 43, 43),
        ("Sprint 22", 48, 45, 45),
        ("Sprint 23", 46, 44, 44),
        ("Sprint 24", 50, 47, 47),
    ]
    
    for sprint_name, committed, completed, velocity in sprints:
        ws.cell(row=row, column=1, value=sprint_name)
        ws.cell(row=row, column=2, value=committed)
        ws.cell(row=row, column=3, value=completed)
        ws.cell(row=row, column=4, value=velocity)
        row += 1
    
    # Add velocity chart
    chart = ChartBuilder.create_velocity_chart(
        ws,
        start_row=data_start_row,
        sprint_col=1,
        committed_col=2,
        completed_col=3,
        num_sprints=5
    )
    ws.add_chart(chart, f"F{data_start_row - 1}")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    
    # Create Priority Analysis sheet
    ws2 = wb.create_sheet("Priority Analysis")
    
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "Priority Distribution"
    ws2['A1'].font = Font(bold=True, size=14, color="1F4E79")
    
    row = 3
    headers = ['Priority', 'Total', 'Open', 'Resolved']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    
    data_start_row = row
    priorities = [
        ("Highest", 12, 3, 9),
        ("High", 35, 8, 27),
        ("Medium", 78, 22, 56),
        ("Low", 45, 15, 30),
    ]
    
    for priority, total, open_count, resolved in priorities:
        ws2.cell(row=row, column=1, value=priority)
        ws2.cell(row=row, column=2, value=total)
        ws2.cell(row=row, column=3, value=open_count)
        ws2.cell(row=row, column=4, value=resolved)
        row += 1
    
    # Add pie chart
    chart = ChartBuilder.create_pie_chart(
        ws2,
        data_range=(3, data_start_row - 1, 3, row - 1),
        categories_range=(1, data_start_row, 1, row - 1),
        title="Open Issues by Priority"
    )
    ws2.add_chart(chart, f"F{data_start_row - 1}")
    
    for col in range(1, 5):
        ws2.column_dimensions[chr(64 + col)].width = 15
    
    # Save workbook
    wb.save(file_path)
    
    print(f"\n{'='*50}")
    print("Demo Report Generated Successfully!")
    print(f"{'='*50}")
    print(f"Output: {file_path.absolute()}")
    print(f"\nThis demonstrates:")
    print("  ✓ Excel dashboard creation")
    print("  ✓ Chart generation (velocity, pie)")
    print("  ✓ Professional formatting")
    print("  ✓ Multiple sheets")
    print(f"\nOpen the file to view the dashboard!")
    
    return str(file_path.absolute())


if __name__ == '__main__':
    try:
        create_demo_report()
    except Exception as e:
        logger.error(f"Demo report failed: {e}")
        print(f"\nError: {e}")
        sys.exit(1)
