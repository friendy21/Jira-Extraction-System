#!/usr/bin/env python
"""Generate one consolidated compliance Excel report for all employees in a provided roster."""

import csv
import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BASE_DIR = Path(__file__).resolve().parent.parent
ROSTER_PATH = BASE_DIR / "data_employee_roster.csv"
AUDIT_JSON_PATH = BASE_DIR / "compliance_report_test.json"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_PATH = OUTPUT_DIR / "JIRA_Consolidated_Employee_Compliance_Report.xlsx"


def load_roster(path: Path):
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def load_audit_map(path: Path):
    if not path.exists():
        return {}

    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    result = {}
    for row in data:
        name = row.get("employee_name")
        if name:
            result[name.strip().lower()] = row
    return result


def val_or_unknown(value):
    return value if value not in (None, "") else "UNKNOWN"


def build_report(roster_rows, audit_map, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Compliance"

    headers = [
        "Employee Name",
        "Email",
        "Role",
        "Account Status",
        "Last Activity",
        "Week Start Date",
        "Status Hygiene",
        "Task Cancellation",
        "Weekly Updates",
        "Roles & Ownership",
        "Documentation/Traceability",
        "Lifecycle Adherence",
        "Zero-Tolerance Violation",
        "Overall Compliance",
        "Auditor Notes",
        "MIT Planning",
        "MIT Creation",
        "MIT Completion",
        "Non-MIT Tracking",
        "Recap-to-Jira Conversion",
        "Comment Quality",
        "Missing Comments",
        "Screenshot-Only Evidence",
        "Doc-Link-Only Evidence",
        "Description Quality",
        "Title Quality",
        "Multiple Issues in One Ticket",
        "History Integrity",
        "Acceptance Criteria Relevance",
        "Productivity Validity",
        "Evidence Relevance",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for idx, emp in enumerate(roster_rows, start=2):
        key = emp["Name"].strip().lower()
        audit = audit_map.get(key, {})

        row = [
            emp.get("Name", ""),
            emp.get("Email", ""),
            val_or_unknown(emp.get("Role", "")),
            val_or_unknown(emp.get("Status", "")),
            val_or_unknown(emp.get("Last Activity", "")),
            val_or_unknown(audit.get("week_start_date")),
            val_or_unknown(audit.get("status_hygiene")),
            val_or_unknown(audit.get("cancellation")),
            val_or_unknown(audit.get("update_frequency")),
            val_or_unknown(audit.get("role_ownership")),
            val_or_unknown(audit.get("documentation")),
            val_or_unknown(audit.get("lifecycle")),
            val_or_unknown(audit.get("zero_tolerance")),
            val_or_unknown(audit.get("overall_compliance")),
            val_or_unknown(audit.get("auditor_notes")),
            "UNKNOWN",
            "UNKNOWN",
            "UNKNOWN",
            "UNKNOWN",
            "UNKNOWN",
            "NEEDS_MANUAL_REVIEW",
            "UNKNOWN",
            "UNKNOWN",
            "UNKNOWN",
            "NEEDS_MANUAL_REVIEW",
            "NEEDS_MANUAL_REVIEW",
            "NEEDS_MANUAL_REVIEW",
            "CRITICAL-UNKNOWN",
            "UNKNOWN",
            "NEEDS_MANUAL_REVIEW",
            "NEEDS_MANUAL_REVIEW",
        ]

        for c, value in enumerate(row, 1):
            cell = ws.cell(row=idx, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    # Basic column widths
    widths = {
        1: 24, 2: 34, 3: 20, 4: 14, 5: 16, 6: 14, 7: 20, 8: 20, 9: 18,
        10: 22, 11: 24, 12: 20, 13: 20, 14: 16, 15: 42
    }
    for col_idx in range(16, len(headers) + 1):
        widths[col_idx] = 22
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    roster = load_roster(ROSTER_PATH)
    audit_map = load_audit_map(AUDIT_JSON_PATH)
    build_report(roster, audit_map, OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")
    print(f"Employees included: {len(roster)}")
    print(f"Generated at: {datetime.now().isoformat(timespec='seconds')}")


if __name__ == "__main__":
    main()
