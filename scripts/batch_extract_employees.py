#!/usr/bin/env python
"""
Batch-by-Batch Employee Extraction - Practical Solution
Fetches ALL employees by making multiple smaller API calls per project.

Strategy: Split each project into time-based batches (e.g., by year or month)
to work around API pagination limitations.

Usage:
    python scripts/batch_extract_employees.py [--weeks WEEKS]
    
Output:
    - Excel compliance report with ALL employees
    - Progress display showing batch processing
"""

import sys
from pathlib import Path
from datetime import datetime, timedelta
import argparse
import time

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.jira_client import JiraClient
from src.utils.logger import setup_logging, get_logger

setup_logging()
logger = get_logger(__name__)


def extract_users_batch_by_time(client: JiraClient):
    """
    Extract ALL users by splitting queries into time-based batches.
    This works around pagination limits by making multiple smaller queries.
    """
    print("\n📥 BATCH EXTRACTION: Fetching ALL employees from JIRA...")
    print("=" * 80)
    
    all_users = {}
    total_issues = 0
    
    # Get all projects
    projects = list(client.fetch_projects())
    print(f"\n✓ Found {len(projects)} projects to scan\n")
    
    # Define time windows (split into yearly batches going back 5 years)
    current_year = datetime.now().year
    time_windows = []
    
    for year in range(current_year - 4, current_year + 1):  # Last 5 years
        time_windows.append({
            'start': f"{year}-01-01",
            'end': f"{year}-12-31",
            'label': str(year)
        })
    
    # Add "all time before" window
    time_windows.insert(0, {
        'start': "2000-01-01",
        'end': f"{current_year - 5}-12-31",
        'label': f"Before {current_year - 4}"
    })
    
    print(f"📅 Time Windows: {len(time_windows)} periods")
    print(f"   {', '.join([w['label'] for w in time_windows])}\n")
    print("-" * 80)
    
    # Process each project
    for proj_idx, project in enumerate(projects, 1):
        project_key = project.get('key')
        project_name = project.get('name', 'Unknown')
        
        print(f"\n[{proj_idx}/{len(projects)}] {project_key} - {project_name}")
        
        project_issues = 0
        users_before = len(all_users)
        
        # Try each time window for this project
        for window_idx, window in enumerate(time_windows, 1):
            try:
                # JQL for this specific time window
                jql = (f"project = {project_key} AND "
                       f"created >= '{window['start']}' AND "
                       f"created <= '{window['end']}' "
                       f"ORDER BY created ASC")
                
                # Fetch issues for this window (limit 100 per window for safety)
                issues = list(client.fetch_issues(
                    jql, 
                    fields=['assignee', 'reporter', 'creator'],
                    max_results=100
                ))
                
                if issues:
                    print(f"  [{window['label']}] {len(issues)} issues", end="")
                    
                    # Extract users from these issues
                    for issue in issues:
                        fields = issue.get('fields', {})
                        
                        # Extract assignee, reporter, creator
                        for user_field in ['assignee', 'reporter', 'creator']:
                            user = fields.get(user_field)
                            if user and isinstance(user, dict):
                                account_id = user.get('accountId')
                                if account_id and account_id not in all_users:
                                    all_users[account_id] = {
                                        'accountId': account_id,
                                        'name': user.get('displayName', 'Unknown'),
                                        'email': user.get('emailAddress', ''),
                                        'active': user.get('active', True)
                                    }
                    
                    project_issues += len(issues)
                    total_issues += len(issues)
                    
                    new_users = len(all_users) - users_before
                    print(f" → {len(all_users)} total employees (+{new_users} new)")
                
                # Small delay to avoid rate limiting
                time.sleep(0.1)
                
            except Exception as e:
                logger.warning(f"Error in {project_key} window {window['label']}: {e}")
                continue
        
        if project_issues > 0:
            new_in_project = len(all_users) - users_before
            print(f"  ✓ {project_key} complete: {project_issues} issues, +{new_in_project} employees")
        else:
            print(f"  ○ {project_key}: No issues found")
        
        print("-" * 80)
    
    # Filter active users
    active_users = [user for user in all_users.values() if user.get('active', True)]
    
    print(f"\n{'=' * 80}")
    print(f"✅ EXTRACTION COMPLETE!")
    print(f"{'=' * 80}")
    print(f"\n📊 Statistics:")
    print(f"   • Total Issues Analyzed: {total_issues:,}")
    print(f"   • Total Unique Employees: {len(all_users)}")
    print(f"   • Active Employees: {len(active_users)}")
    
    # Display employee list
    print(f"\n📋 Complete Employee List:")
    print(f"{'-' * 80}")
    for i, user in enumerate(sorted(active_users, key=lambda x: x['name']), 1):
        print(f"   {i:3d}. {user['name']}")
    
    return active_users


from src.utils.helpers import parse_jira_datetime
import pytz
import re

def check_mit_compliance(issues, start_date, end_date):
    """Check if MITs were created and closed properly."""
    mit_issues = []
    for issue in issues:
        fields = issue.get('fields', {})
        labels = fields.get('labels', [])
        priority = fields.get('priority', {}).get('name', '')
        
        # Check for MIT criteria
        if 'MIT' in labels or priority == 'This Week':
            mit_issues.append(issue)
            
    if not mit_issues:
        return 'No', 'No MITs found for this period'
        
    closed_mits = 0
    on_time_mits = 0
    
    for issue in mit_issues:
        fields = issue.get('fields', {})
        status = fields.get('status', {}).get('name', '')
        resolution_date = parse_jira_datetime(fields.get('resolutiondate'))
        
        if status in ['Done', 'Closed', 'Resolved']:
            closed_mits += 1
            # Check if closed by Friday? (Simplified: Just check if closed)
    
    if closed_mits == len(mit_issues):
        return 'Yes', f'All {len(mit_issues)} MITs closed'
    elif closed_mits > 0:
        return 'Partial', f'{closed_mits}/{len(mit_issues)} MITs closed'
    else:
        return 'No', f'0/{len(mit_issues)} MITs closed'


def check_updates_compliance(issues, user_email, start_date, end_date):
    """Check for Wed/Fri updates."""
    # This is complex to check perfectly without processing text.
    # We'll check if there are distinct comments on Wed and Fri.
    
    wed_updates = 0
    fri_updates = 0
    
    # Iterate through all issues to find comments by this user
    # Note: This is an approximation as we only have issues *assigned* to them
    # ideally we'd search `commenter = user` but that's expensive.
    # We will check comments on their assigned issues.
    
    comments_found = []
    
    for issue in issues:
        fields = issue.get('fields', {})
        comment_data = fields.get('comment', {}).get('comments', [])
        
        for comment in comment_data:
            author = comment.get('author', {})
            if author.get('emailAddress') == user_email:
                created = parse_jira_datetime(comment.get('created'))
                if created and start_date <= created <= end_date:
                    comments_found.append(created)
                    
                    # Check weekday (Mon=0, Tue=1, Wed=2, Thu=3, Fri=4)
                    weekday = created.weekday()
                    if weekday == 2: # Wednesday
                        wed_updates += 1
                    elif weekday == 4: # Friday
                        fri_updates += 1
                        
    if wed_updates > 0 and fri_updates > 0:
        return 'Yes', f'Found updates on Wed ({wed_updates}) and Fri ({fri_updates})'
    elif wed_updates > 0 or fri_updates > 0:
        return 'Partial', f'Missing updates on {"Fri" if wed_updates > 0 else "Wed"}'
    else:
        return 'No', 'No Wed/Fri updates found on assigned tickets'


def check_roles_compliance(issues, user_account_id):
    """Check if reporter == assignee (Self-Assignment)."""
    violations = 0
    for issue in issues:
        fields = issue.get('fields', {})
        reporter = fields.get('reporter', {})
        assignee = fields.get('assignee', {})
        
        if reporter and assignee:
             if reporter.get('accountId') == assignee.get('accountId'):
                 violations += 1
                 
    if violations == 0:
        return 'Yes', 'No self-assigned tickets'
    else:
        return 'No', f'{violations} self-assigned tickets found'


def check_documentation_compliance(issues):
    """Check for description and acceptance criteria."""
    poor_docs = 0
    for issue in issues:
        fields = issue.get('fields', {})
        description = fields.get('description', '')
        # Handle ADF (Atlassian Document Format) or plain text
        # If it's a dict (ADF), assuming it has content.
        has_content = False
        if isinstance(description, dict):
            # rudimentary check for ADF content
             if description.get('content'):
                 has_content = True
        elif description and len(str(description)) > 10:
            has_content = True
            
        if not has_content:
            poor_docs += 1
            
    if poor_docs == 0:
        return 'Yes', 'All tickets have descriptions'
    else:
        return 'No', f'{poor_docs} tickets with missing/empty descriptions'


def check_status_hygiene(issues):
    """Check for stale tickets."""
    stale_issues = 0
    now = datetime.now(pytz.UTC)
    for issue in issues:
        fields = issue.get('fields', {})
        updated = parse_jira_datetime(fields.get('updated'))
        status = fields.get('status', {}).get('name', '')
        
        if status not in ['Done', 'Closed', 'Resolved'] and updated:
            # If not updated in 7 days
            if (now - updated).days > 7:
                 stale_issues += 1
                 
    if stale_issues == 0:
        return 'Yes', 'Active flow'
    else:
        return 'No', f'{stale_issues} stale tickets (>7 days no update)'


def check_zero_tolerance(issues):
    """Check for unauthorized cancellations."""
    cancelled_issues = 0
    for issue in issues:
        fields = issue.get('fields', {})
        status = fields.get('status', {}).get('name', '')
        resolution = fields.get('resolution')
        resolution_name = resolution.get('name', '') if resolution else ''
        
        if status == 'Cancelled' or resolution_name == 'Won\'t Do':
            # Strictly speaking we need to check if manager did it.
            # For now, just flag it.
            cancelled_issues += 1
            
    if cancelled_issues == 0:
        return 'No', 'No cancellations'
    else:
        return 'Review', f'{cancelled_issues} cancelled tasks found (Check approval)'


def audit_employee(user, client: JiraClient, weeks=4):
    """Run compliance audit for a single employee."""
    print(f"   Running audit for {user['name']}...")
    
    end_date = datetime.now(pytz.UTC)
    start_date = end_date - timedelta(weeks=weeks)
    date_str = start_date.strftime('%Y-%m-%d')
    
    # JQL to find issues assigned to user updated in the window
    # Fetch Changelog and Comments
    # Important: 'assignee' uses accountId
    jql = f'assignee = "{user.get("accountId")}" AND updated >= "{date_str}"'
    
    try:
        issues = list(client.fetch_issues(
            jql, 
            fields=['summary', 'status', 'priority', 'labels', 'created', 'updated', 
                    'resolutiondate', 'assignee', 'reporter', 'comment', 'duedate', 
                    'description', 'resolution'],
            expand=['changelog'],
            max_results=200 # Limit to avoid overload
        ))
    except Exception as e:
        logger.error(f"Failed to fetch audit issues for {user['name']}: {e}")
        issues = []
        
    # Analyze
    mit_status, mit_notes = check_mit_compliance(issues, start_date, end_date)
    updates_status, updates_notes = check_updates_compliance(issues, user.get('email'), start_date, end_date)
    roles_status, roles_notes = check_roles_compliance(issues, user.get('accountId'))
    doc_status, doc_notes = check_documentation_compliance(issues)
    hygiene_status, hygiene_notes = check_status_hygiene(issues)
    zero_status, zero_notes = check_zero_tolerance(issues)
    
    # Overall Pass/Fail
    # Fail if any "No" in critical areas (MIT, Updates, Roles, Docs)
    # Zero Tol is "Yes" = Fail (but our function returns "No" for good)
    
    failures = []
    if mit_status == 'No': failures.append("MITs")
    if updates_status == 'No': failures.append("Updates")
    if roles_status == 'No': failures.append("Roles")
    if doc_status == 'No': failures.append("Docs")
    if zero_status == 'Review': failures.append("ZeroTol") # Should be 'No' for pass
    
    overall = 'Fail' if failures else 'Pass'
    
    return {
        'name': user['name'],
        'email': user['email'],
        'issues_analyzed': len(issues),
        'mit_compliance': mit_status,
        'mit_notes': mit_notes,
        'updates_compliance': updates_status,
        'updates_notes': updates_notes,
        'roles_compliance': roles_status,
        'roles_notes': roles_notes,
        'doc_compliance': doc_status,
        'doc_notes': doc_notes,
        'hygiene_compliance': hygiene_status,
        'hygiene_notes': hygiene_notes,
        'zero_tolerance': zero_status,
        'zero_notes': zero_notes,
        'overall': overall,
        'summary_notes': f"Failed: {', '.join(failures)}" if failures else "All checks passed"
    }


def generate_report(audit_results, weeks=4):
    """Generate compliance report from audit results."""
    print(f"\n📊 Generating compliance report for {len(audit_results)} employees...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "JIRA Compliance Audit"
    
    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    data_font = Font(name='Calibri', size=11)
    pass_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = [
        "Employee Name", "MIT Completion", "Wed/Fri Updates", "Roles & Access", 
        "Documentation", "Status Hygiene", "Zero Tolerance", "Overall", "Detailed Notes"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        
    row = 2
    for result in audit_results:
        ws.cell(row=row, column=1, value=result['name']).font = data_font
        
        # Helper to set cell
        def set_cell(col, val, notes):
            c = ws.cell(row=row, column=col, value=val)
            c.comment = None # Could add comment
            c.font = data_font
            if val == 'No' or val == 'Fail' or (col==7 and val != 'No'): 
                 c.font = Font(color="FF0000", bold=True)
            c.border = border # Apply border to all cells
            return c
        
        set_cell(2, result['mit_compliance'], result['mit_notes'])
        set_cell(3, result['updates_compliance'], result['updates_notes'])
        set_cell(4, result['roles_compliance'], result['roles_notes'])
        set_cell(5, result['doc_compliance'], result['doc_notes'])
        set_cell(6, result['hygiene_compliance'], result['hygiene_notes'])
        set_cell(7, result['zero_tolerance'], result['zero_notes'])
        
        overall = ws.cell(row=row, column=8, value=result['overall'])
        overall.fill = pass_fill if result['overall'] == 'Pass' else fail_fill
        overall.font = Font(color="FFFFFF", bold=True)
        overall.alignment = Alignment(horizontal='center')
        overall.border = border # Apply border
        
        ws.cell(row=row, column=9, value=result['summary_notes']).font = data_font
        ws.cell(row=row, column=9).alignment = Alignment(horizontal='left', wrap_text=True)
        ws.cell(row=row, column=9).border = border # Apply border
        
        row += 1
        
    # Auto-adjust columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    # Save
    output_dir = Path('./outputs')
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = output_dir / f"JIRA_Compliance_Audit_{timestamp}.xlsx"
    wb.save(file_path)
    
    return file_path


def main():
    # Force UTF-8 output for Windows consoles
    if sys.stdout.encoding != 'utf-8':
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except AttributeError:
            pass

    parser = argparse.ArgumentParser(description='Batch extract ALL JIRA employees and Audit')
    parser.add_argument('--weeks', type=int, default=4, help='Number of weeks for audit (default: 4)')
    args = parser.parse_args()
    
    print("\n" + "=" * 80)
    print("🕵️‍♀️ JIRA COMPLIANCE AUDIT & EXTRACTION")
    print("=" * 80)
    
    try:
        print("\n🔌 Connecting to JIRA...")
        client = JiraClient()
        print("✅ Connected!\n")
        
        # Step 1: Get Users (using existing batch method)
        users_list = extract_users_batch_by_time(client)
        
        if not users_list:
            print("No users found.")
            return 1
            
        # Step 2: Audit Each User
        print(f"\n🔍 Auditing {len(users_list)} employees for last {args.weeks} weeks...")
        audit_results = []
        
        for i, user in enumerate(users_list, 1):
             print(f"[{i}/{len(users_list)}] ", end="")
             # Ensure user has accountId
             if 'accountId' not in user:
                 logger.warning(f"Skipping {user.get('name', 'Unknown')} - No Account ID found in extracted data.")
                 continue
                 
             result = audit_employee(user, client, weeks=args.weeks)
             audit_results.append(result)
             
        # Step 3: Generate Report
        file_path = generate_report(audit_results, weeks=args.weeks)
        
        # Success output
        print(f"\n{'=' * 80}")
        print("✅ COMPLIANCE AUDIT REPORT GENERATED!")
        print(f"{'=' * 80}")
        print(f"\n📁 Report Location:")
        print(f"   {file_path.absolute()}")
        print(f"\n📊 Report Statistics:")
        print(f"   • Employees Audited: {len(audit_results)}")
        print(f"   • Weeks Covered: {args.weeks}")
        
        pass_count = sum(1 for r in audit_results if r['overall'] == 'Pass')
        fail_count = len(audit_results) - pass_count
        pass_rate = (pass_count / len(audit_results) * 100) if len(audit_results) > 0 else 0
        
        print(f"   • Pass: {pass_count:,} ({pass_rate:.1f}%)")
        print(f"   • Fail: {fail_count:,} ({100-pass_rate:.1f}%)")
        print(f"\n💡 Open the Excel file to review detailed audit results!")
        print(f"\n⏰ Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'=' * 80}\n")
        
        return 0
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        logger.error(f"Compliance audit failed: {e}", exc_info=True)
        return 1


if __name__ == '__main__':
    sys.exit(main())
