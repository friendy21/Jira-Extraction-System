#!/usr/bin/env python
"""
Demo Compliance Report Generator
Creates sample JIRA compliance report with mock data for demonstration and validation.

Usage:
    python scripts/demo_compliance_report.py
    
Output:
    ./outputs/JIRA_Compliance_Report_YYYYMMDD_HHMMSS.xlsx
"""

import sys
from pathlib import Path
from datetime import datetime, timedelta

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.logger import setup_logging, get_logger

setup_logging()
logger = get_logger(__name__)


def create_demo_compliance_report():
    """Create demo compliance report with mock data."""
    
    # Ensure output directory exists
    output_path = Path('./outputs')
    output_path.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    file_path = output_path / f"JIRA_Compliance_Report_DEMO_{timestamp}.xlsx"
    
    logger.info("Creating demo compliance report...")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "JIRA Compliance Report"
    
    # Define styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    data_font = Font(name='Calibri', size=11)
    pass_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    fail_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Define column headers (exactly as specified)
    headers = [
        "Employee Name",
        "Week Start Date",
        "Status Hygiene Correct",
        "Any Tasks Cancelled w/o Approval",
        "Wed/Fri Updates Shared",
        "Roles & Ownership Correct",
        "Documentation & Traceability Complete",
        "Lifecycle Adherence",
        "Zero-Tolerance Violation",
        "Overall Compliance (Pass/Fail)",
        "Auditor's Notes"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Create mock data (realistic scenarios) - 30 employees across 3 weeks
    # Week starts on Monday (ISO week)
    week1 = datetime(2026, 1, 26)  # Monday, Jan 26, 2026
    week2 = datetime(2026, 2, 2)   # Monday, Feb 2, 2026
    week3 = datetime(2026, 2, 9)   # Monday, Feb 9, 2026
    
    # Employee list (30 employees)
    employees = [
        'Gaurav Dadheech', 'Priya Sharma', 'Rahul Kumar', 'Aarti Singh', 'Amit Patel',
        'Sneha Reddy', 'Vikram Malhotra', 'Neha Gupta', 'Arjun Verma', 'Kavita Nair',
        'Rohan Desai', 'Ananya Iyer', 'Sanjay Chopra', 'Divya Menon', 'Karthik Rao',
        'Megha Bose', 'Nikhil Sinha', 'Pooja Joshi', 'Tarun Singh', 'Ritu Kapoor',
        'Varun Agarwal', 'Simran Kaur', 'Deepak Mehta', 'Swati Bansal', 'Aditya Sharma',
        'Ishita Das', 'Manish Tiwari', 'Preeti Saxena', 'Rajesh Pandey', 'Anjali Mishra'
    ]
    
    # Compliance scenarios (realistic mix)
    scenarios = {
        'perfect': {
            'status_hygiene': 'Yes', 'cancellation': 'No', 'updates': 'Yes',
            'roles': 'Yes', 'documentation': 'Yes', 'lifecycle': 'Yes',
            'zero_tolerance': 'No', 'overall': 'Pass', 'notes': 'All checks passed'
        },
        'missing_updates': {
            'status_hygiene': 'Yes', 'cancellation': 'No', 'updates': 'No',
            'roles': 'Yes', 'documentation': 'Yes', 'lifecycle': 'Yes',
            'zero_tolerance': 'No', 'overall': 'Fail', 'notes': 'Missing Wed/Fri status updates'
        },
        'partial_updates': {
            'status_hygiene': 'Yes', 'cancellation': 'No', 'updates': 'Partial',
            'roles': 'Yes', 'documentation': 'Yes', 'lifecycle': 'Yes',
            'zero_tolerance': 'No', 'overall': 'Fail', 'notes': 'Only Friday update provided, missing Wednesday'
        },
        'role_issues': {
            'status_hygiene': 'Yes', 'cancellation': 'No', 'updates': 'Yes',
            'roles': 'No - Reporter = Assignee', 'documentation': 'Yes', 'lifecycle': 'Yes',
            'zero_tolerance': 'No', 'overall': 'Fail', 'notes': 'Reporter and Assignee are the same person'
        },
        'doc_issues': {
            'status_hygiene': 'Yes', 'cancellation': 'No', 'updates': 'Yes',
            'roles': 'Yes', 'documentation': 'No - No traceability links', 'lifecycle': 'Yes',
            'zero_tolerance': 'No', 'overall': 'Fail', 'notes': 'Missing linked issues for traceability'
        },
        'status_hygiene_fail': {
            'status_hygiene': 'No - Invalid transition', 'cancellation': 'No', 'updates': 'Yes',
            'roles': 'Yes', 'documentation': 'Yes', 'lifecycle': 'Yes',
            'zero_tolerance': 'No', 'overall': 'Fail', 'notes': 'Invalid workflow transition detected'
        },
        'lifecycle_violation': {
            'status_hygiene': 'Yes', 'cancellation': 'No', 'updates': 'Yes',
            'roles': 'Yes', 'documentation': 'Yes', 'lifecycle': 'No - Skipped In Progress',
            'zero_tolerance': 'No', 'overall': 'Fail', 'notes': 'Lifecycle violation: skipped In Progress status'
        },
        'unauthorized_cancel': {
            'status_hygiene': 'Yes', 'cancellation': 'Yes - Task cancelled w/o approval', 'updates': 'Yes',
            'roles': 'Yes', 'documentation': 'Yes', 'lifecycle': 'Yes',
            'zero_tolerance': 'No', 'overall': 'Fail', 'notes': 'Task cancelled without manager approval'
        },
        'multiple_violations': {
            'status_hygiene': 'No', 'cancellation': 'No', 'updates': 'No',
            'roles': 'No - Assignee missing', 'documentation': 'No - Description incomplete',
            'lifecycle': 'No - Skipped In Progress', 'zero_tolerance': 'Yes',
            'overall': 'Fail', 'notes': 'Multiple violations; Requires immediate training'
        },
        'minor_issues': {
            'status_hygiene': 'Yes', 'cancellation': 'No', 'updates': 'Yes',
            'roles': 'Yes', 'documentation': 'No - Due date missing', 'lifecycle': 'Yes',
            'zero_tolerance': 'No', 'overall': 'Fail', 'notes': 'Due date not set on tasks'
        }
    }
    
    mock_data = []
    
    # Week 1: Generate varied compliance data
    scenario_keys = list(scenarios.keys())
    for i, employee in enumerate(employees):
        # Distribute scenarios across employees (weighted toward pass/minor issues)
        if i % 10 == 0:
            scenario = scenarios['perfect']
        elif i % 10 == 1:
            scenario = scenarios['missing_updates']
        elif i % 10 == 2:
            scenario = scenarios['role_issues']
        elif i % 10 == 3:
            scenario = scenarios['doc_issues']
        elif i % 10 == 4:
            scenario = scenarios['perfect']
        elif i % 10 == 5:
            scenario = scenarios['partial_updates']
        elif i % 10 == 6:
            scenario = scenarios['status_hygiene_fail']
        elif i % 10 == 7:
            scenario = scenarios['perfect']
        elif i % 10 == 8:
            scenario = scenarios['minor_issues']
        else:
            scenario = scenarios['lifecycle_violation']
        
        mock_data.append({
            'employee': employee,
            'week': week1,
            **scenario
        })
    
    # Week 2: Show improvements for some, new issues for others
    for i, employee in enumerate(employees):
        if i % 10 == 0:
            scenario = scenarios['perfect']
        elif i % 10 == 1:
            scenario = scenarios['perfect']  # Improved from week 1
        elif i % 10 == 2:
            scenario = scenarios['partial_updates']  # Still struggling
        elif i % 10 == 3:
            scenario = scenarios['perfect']  # Improved
        elif i % 10 == 4:
            scenario = scenarios['perfect']
        elif i % 10 == 5:
            scenario = scenarios['unauthorized_cancel']  # New issue
        elif i % 10 == 6:
            scenario = scenarios['minor_issues']  # Improved
        elif i % 10 == 7:
            scenario = scenarios['perfect']
        elif i % 10 == 8:
            scenario = scenarios['missing_updates']  # Regressed
        else:
            scenario = scenarios['doc_issues']
        
        mock_data.append({
            'employee': employee,
            'week': week2,
            **scenario
        })
    
    # Week 3: Further improvements, some persistent issues
    for i, employee in enumerate(employees):
        if i % 10 == 0:
            scenario = scenarios['perfect']
        elif i % 10 == 1:
            scenario = scenarios['perfect']  # Maintained improvement
        elif i % 10 == 2:
            scenario = scenarios['perfect']  # Finally improved
        elif i % 10 == 3:
            scenario = scenarios['perfect']
        elif i % 10 == 4:
            scenario = scenarios['perfect']
        elif i % 10 == 5:
            scenario = scenarios['role_issues']  # Different issue
        elif i % 10 == 6:
            scenario = scenarios['perfect']  # Maintained
        elif i % 10 == 7:
            scenario = scenarios['perfect']
        elif i % 10 == 8:
            scenario = scenarios['perfect']  # Improved
        else:
            scenario = scenarios['multiple_violations']  # Persistent problems
        
        mock_data.append({
            'employee': employee,
            'week': week3,
            **scenario
        })
    
    
    # Write data rows
    for row_idx, data in enumerate(mock_data, 2):
        # Column A: Employee Name
        cell = ws.cell(row=row_idx, column=1, value=data['employee'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Column B: Week Start Date
        date_cell = ws.cell(row=row_idx, column=2, value=data['week'])
        date_cell.number_format = 'YYYY-MM-DD'
        date_cell.font = data_font
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.border = border
        
        # Column C: Status Hygiene Correct
        cell = ws.cell(row=row_idx, column=3, value=data['status_hygiene'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Column D: Any Tasks Cancelled w/o Approval
        cell = ws.cell(row=row_idx, column=4, value=data['cancellation'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Column E: Wed/Fri Updates Shared
        cell = ws.cell(row=row_idx, column=5, value=data['updates'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Column F: Roles & Ownership Correct
        cell = ws.cell(row=row_idx, column=6, value=data['roles'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Column G: Documentation & Traceability Complete
        cell = ws.cell(row=row_idx, column=7, value=data['documentation'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Column H: Lifecycle Adherence
        cell = ws.cell(row=row_idx, column=8, value=data['lifecycle'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Column I: Zero-Tolerance Violation
        cell = ws.cell(row=row_idx, column=9, value=data['zero_tolerance'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left')
        cell.border = border
        
        # Column J: Overall Compliance (Pass/Fail)
        overall_cell = ws.cell(row=row_idx, column=10, value=data['overall'])
        overall_cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        overall_cell.alignment = Alignment(horizontal='center')
        overall_cell.border = border
        
        # Apply conditional formatting (green for Pass, red for Fail)
        if data['overall'] == 'Pass':
            overall_cell.fill = pass_fill
        else:
            overall_cell.fill = fail_fill
        
        # Column K: Auditor's Notes
        notes_cell = ws.cell(row=row_idx, column=11, value=data['notes'])
        notes_cell.font = data_font
        notes_cell.alignment = Alignment(horizontal='left', wrap_text=True)
        notes_cell.border = border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Employee Name
    ws.column_dimensions['B'].width = 15  # Week Start Date
    ws.column_dimensions['C'].width = 22  # Status Hygiene
    ws.column_dimensions['D'].width = 28  # Cancellation
    ws.column_dimensions['E'].width = 20  # Updates
    ws.column_dimensions['F'].width = 28  # Roles
    ws.column_dimensions['G'].width = 32  # Documentation
    ws.column_dimensions['H'].width = 18  # Lifecycle
    ws.column_dimensions['I'].width = 22  # Zero-Tolerance
    ws.column_dimensions['J'].width = 22  # Overall Compliance
    ws.column_dimensions['K'].width = 50  # Auditor's Notes
    
    # Save workbook
    wb.save(file_path)
    
    # Print success message
    print(f"\n{'='*70}")
    print("✅ Demo Compliance Report Generated Successfully!")
    print(f"{'='*70}")
    print(f"📁 Output: {file_path.absolute()}")
    print(f"\n📊 Report Structure:")
    print(f"   ✓ Single sheet: 'JIRA Compliance Report'")
    print(f"   ✓ 11 columns (A-K)")
    print(f"   ✓ {len(mock_data)} employee-week compliance records")
    print(f"   ✓ Pass/Fail indicators with color coding")
    print(f"   ✓ Detailed auditor notes")
    print(f"\n📈 Sample Data Summary:")
    print(f"   - Employees: {len(set(d['employee'] for d in mock_data))} total")
    print(f"   - Weeks: {week1.strftime('%Y-%m-%d')}, {week2.strftime('%Y-%m-%d')}, {week3.strftime('%Y-%m-%d')}")
    print(f"   - Total Records: {len(mock_data)}")
    print(f"   - Pass Rate: {sum(1 for d in mock_data if d['overall'] == 'Pass')}/{len(mock_data)} ({100 * sum(1 for d in mock_data if d['overall'] == 'Pass') / len(mock_data):.1f}%)")
    print(f"\n💡 Open the file to view the compliance report!")
    print(f"{'='*70}\n")
    
    logger.info(f"Demo compliance report created: {file_path}")
    
    return str(file_path.absolute())


if __name__ == '__main__':
    try:
        create_demo_compliance_report()
    except Exception as e:
        logger.error(f"Demo report generation failed: {e}", exc_info=True)
        print(f"\n❌ Error: {e}")
        sys.exit(1)
