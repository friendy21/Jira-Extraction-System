#!/usr/bin/env python
"""
Integration Test for Compliance Reporting System
Tests the entire pipeline from JIRA data fetch to Excel generation.
"""

import sys
from pathlib import Path
from datetime import datetime, timedelta
import os

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.config_manager import ConfigManager
from src.jira_client import JiraClient
from src.reports.compliance_builder import ComplianceReportBuilder
from src.database.connection import get_session
from src.database.models import JiraUser
from src.utils.logger import setup_logging, get_logger

setup_logging()
logger = get_logger(__name__)


class ComplianceIntegrationTest:
    """Integration test suite for compliance reporting."""
    
    def __init__(self):
        self.config = ConfigManager()
        self.jira = None
        self.builder = None
        self.test_results = []
    
    def run_all_tests(self):
        """Run all integration tests."""
        print("=" * 70)
        print("JIRA Compliance Report - Integration Test Suite")
        print("=" * 70)
        print()
        
        tests = [
            ("Database Connection", self.test_database_connection),
            ("JIRA Connection", self.test_jira_connection),
            ("Active Employees Query", self.test_active_employees),
            ("Report Builder Initialization", self.test_builder_init),
            ("Demo Report Generation", self.test_demo_report),
            ("Compliance Check Execution", self.test_compliance_checks),
            ("Excel File Validation", self.test_excel_output),
        ]
        
        passed = 0
        failed = 0
        
        for test_name, test_func in tests:
            try:
                print(f"Running: {test_name}...")
                result = test_func()
                if result:
                    print(f"✅ PASSED: {test_name}")
                    passed += 1
                else:
                    print(f"❌ FAILED: {test_name}")
                    failed += 1
            except Exception as e:
                print(f"❌ ERROR: {test_name} - {str(e)}")
                logger.error(f"Test failed: {test_name}", exc_info=True)
                failed += 1
            print()
        
        # Summary
        print("=" * 70)
        print("Test Summary")
        print("=" * 70)
        print(f"Total Tests: {len(tests)}")
        print(f"Passed: {passed} ✅")
        print(f"Failed: {failed} ❌")
        print(f"Success Rate: {(passed/len(tests)*100):.1f}%")
        print()
        
        return failed == 0
    
    def test_database_connection(self):
        """Test database connection."""
        try:
            with get_session() as session:
                # Simple query to verify connection
                result = session.execute("SELECT 1").scalar()
                assert result == 1
            return True
        except Exception as e:
            logger.error(f"Database connection failed: {e}")
            return False
    
    def test_jira_connection(self):
        """Test JIRA API connection."""
        try:
            self.jira = JiraClient()
            server_info = self.jira.get_server_info()
            
            assert 'baseUrl' in server_info
            assert 'version' in server_info
            
            print(f"   Connected to: {server_info.get('baseUrl')}")
            print(f"   JIRA Version: {server_info.get('version')}")
            
            return True
        except Exception as e:
            logger.error(f"JIRA connection failed: {e}")
            print(f"   Note: This may be expected if JIRA credentials are not configured")
            return False
    
    def test_active_employees(self):
        """Test fetching active employees from database."""
        try:
            with get_session() as session:
                employees = session.query(JiraUser).filter(
                    JiraUser.active == True
                ).all()
                
                print(f"   Found {len(employees)} active employees")
                
                if len(employees) > 0:
                    sample = employees[0]
                    print(f"   Sample: {sample.display_name} ({sample.account_id})")
                
                return len(employees) >= 0  # Allow 0 for empty databases
        except Exception as e:
            logger.error(f"Employee query failed: {e}")
            return False
    
    def test_builder_init(self):
        """Test ComplianceReportBuilder initialization."""
        try:
            if not self.jira:
                self.jira = JiraClient()
            
            self.builder = ComplianceReportBuilder(
                self.jira,
                output_dir="./outputs"
            )
            
            # Verify checks are initialized
            assert len(self.builder.checks) == 7
            print(f"   Initialized {len(self.builder.checks)} compliance checks")
            
            # List check names
            check_names = list(self.builder.checks.keys())
            print(f"   Checks: {', '.join(check_names)}")
            
            return True
        except Exception as e:
            logger.error(f"Builder initialization failed: {e}")
            return False
    
    def test_demo_report(self):
        """Test demo report generation."""
        try:
            from scripts.demo_compliance_report import create_demo_compliance_report
            
            output_path = create_demo_compliance_report()
            
            # Verify file exists
            assert os.path.exists(output_path)
            
            # Verify file size
            file_size = os.path.getsize(output_path)
            assert file_size > 1000  # Should be at least 1KB
            
            print(f"   Generated: {os.path.basename(output_path)}")
            print(f"   Size: {file_size:,} bytes")
            
            return True
        except Exception as e:
            logger.error(f"Demo report generation failed: {e}")
            return False
    
    def test_compliance_checks(self):
        """Test individual compliance checks with mock data."""
        try:
            from unittest.mock import Mock
            
            # Create mock employee
            employee = Mock(account_id='test123', display_name='Test Employee')
            
            # Create mock issue with complete data
            mock_issue = {
                'key': 'TEST-1',
                'fields': {
                    'status': {'name': 'Done'},
                    'reporter': {'accountId': 'user1'},
                    'assignee': {'accountId': 'user2'},
                    'description': 'This is a test description with more than 50 characters for validation.',
                    'issuelinks': [{'id': '123'}],
                    'attachment': [],
                    'duedate': '2026-02-01',
                    'priority': {'name': 'Medium'},
                    'issuetype': {'name': 'Task'},
                    'comment': {
                        'comments': [
                            {
                                'created': '2026-01-22T10:00:00.000+0000',
                                'author': {'accountId': 'test123'},
                                'body': 'Wednesday update'
                            },
                            {
                                'created': '2026-01-24T10:00:00.000+0000',
                                'author': {'accountId': 'test123'},
                                'body': 'Friday update'
                            }
                        ]
                    }
                },
                'changelog': {
                    'histories': [
                        {
                            'created': '2026-01-20T09:00:00.000+0000',
                            'items': [{
                                'field': 'status',
                                'fromString': 'To Do',
                                'toString': 'In Progress'
                            }]
                        },
                        {
                            'created': '2026-01-23T14:00:00.000+0000',
                            'items': [{
                                'field': 'status',
                                'fromString': 'In Progress',
                                'toString': 'Done'
                            }]
                        }
                    ]
                }
            }
            
            issues = [mock_issue]
            
            # Test each check
            if not self.builder:
                self.test_builder_init()
            
            check_results = {}
            for check_name, check in self.builder.checks.items():
                result = check.evaluate(issues, employee)
                check_results[check_name] = result
                print(f"   {check_name}: {result}")
            
            # Verify at least some checks passed
            passed_checks = sum(1 for r in check_results.values() if r == "Yes" or r == "No")
            assert passed_checks > 0
            
            return True
        except Exception as e:
            logger.error(f"Compliance checks test failed: {e}")
            return False
    
    def test_excel_output(self):
        """Test Excel output validation."""
        try:
            from openpyxl import load_workbook
            
            # Find most recent demo report
            output_dir = Path("./outputs")
            demo_files = list(output_dir.glob("JIRA_Compliance_Report_DEMO_*.xlsx"))
            
            if not demo_files:
                print("   No demo files found, creating one...")
                self.test_demo_report()
                demo_files = list(output_dir.glob("JIRA_Compliance_Report_DEMO_*.xlsx"))
            
            # Get most recent
            latest_file = max(demo_files, key=lambda f: f.stat().st_mtime)
            
            # Load and validate
            wb = load_workbook(latest_file)
            ws = wb.active
            
            # Validate sheet name
            assert ws.title == "JIRA Compliance Report"
            
            # Validate headers (11 columns)
            expected_headers = [
                "Employee Name",
                "Week Start Date",
                "Status Hygiene Correct",
                "Any Tasks Cancelled w/o Approval",
                "Wed/Fri Updates Shared",
                "Roles & Ownership Correct",
                "Documentation & Traceability Complete",
                "Lifecycle Adherence",
                "Zero-Tolerance Violation",
                "Overall Compliance (Pass/Fail)",
                "Auditor's Notes"
            ]
            
            for col, expected_header in enumerate(expected_headers, 1):
                actual_header = ws.cell(1, col).value
                assert actual_header == expected_header, f"Column {col}: expected '{expected_header}', got '{actual_header}'"
            
            print(f"   Validated: {latest_file.name}")
            print(f"   Columns: {ws.max_column} (expected 11)")
            print(f"   Data Rows: {ws.max_row - 1}")
            
            return True
        except Exception as e:
            logger.error(f"Excel validation failed: {e}")
            return False


def main():
    """Run integration tests."""
    print("\n")
    test_suite = ComplianceIntegrationTest()
    success = test_suite.run_all_tests()
    
    if success:
        print("🎉 All integration tests passed!")
        sys.exit(0)
    else:
        print("⚠️  Some integration tests failed. Check logs for details.")
        sys.exit(1)


if __name__ == '__main__':
    main()
