"""
Compliance Report Builder Module
Generates JIRA compliance audit reports tracking employee weekly process adherence.
"""

from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.compliance.checks import (
    StatusHygieneCheck,
    CancellationCheck,
    UpdateFrequencyCheck,
    RoleOwnershipCheck,
    DocumentationCheck,
    LifecycleCheck,
    ZeroToleranceCheck
)
from src.jira_client import JiraClient
from src.database.connection import get_session
from src.database.models import JiraUser
from src.database.queries import QueryHelpers
from src.utils.logger import get_logger

logger = get_logger(__name__)


class ComplianceReportBuilder:
    """
    Builds JIRA Compliance Reports tracking employee weekly process adherence.
    
    Generates single-sheet Excel reports with 11 columns:
    - Employee Name
    - Week Start Date
    - Status Hygiene Correct
    - Any Tasks Cancelled w/o Approval
    - Wed/Fri Updates Shared
    - Roles & Ownership Correct
    - Documentation & Traceability Complete
    - Lifecycle Adherence
    - Zero-Tolerance Violation
    - Overall Compliance (Pass/Fail)
    - Auditor's Notes
    """
    
    # Column definitions (exact order and naming)
    COLUMNS = [
        "Employee Name",
        "Week Start Date",
        "Status Hygiene Correct",
        "Any Tasks Cancelled w/o Approval",
        "Wed/Fri Updates Shared",
        "Roles & Ownership Correct",
        "Documentation & Traceability Complete",
        "Lifecycle Adherence",
        "Zero-Tolerance Violation",
        "Overall Compliance (Pass/Fail)",
        "Auditor's Notes"
    ]
    
    # Styling constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    DATA_FONT = Font(name='Calibri', size=11)
    PASS_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    def __init__(self, jira_client: JiraClient, output_dir: str = "./outputs"):
        """
        Initialize compliance report builder.
        
        Args:
            jira_client: Authenticated JIRA client instance
            output_dir: Directory for output files
        """
        self.jira = jira_client
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize compliance checks
        self.checks = {
            'status_hygiene': StatusHygieneCheck(),
            'cancellation': CancellationCheck(),
            'update_frequency': UpdateFrequencyCheck(),
            'role_ownership': RoleOwnershipCheck(),
            'documentation': DocumentationCheck(),
            'lifecycle': LifecycleCheck(),
            'zero_tolerance': ZeroToleranceCheck()
        }
        
        logger.info("Compliance report builder initialized")
    
    def generate_report(
        self,
        start_date: datetime,
        end_date: datetime,
        team_id: Optional[int] = None
    ) -> str:
        """
        Generate compliance report for date range.
        
        Args:
            start_date: Report start date
            end_date: Report end date
            team_id: Optional team filter
            
        Returns:
            Path to generated Excel file
        """
        logger.info(f"Generating compliance report: {start_date.date()} to {end_date.date()}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "JIRA Compliance Report"
        
        # Write headers
        self._write_headers(ws)
        
        # Get employees and weeks
        employees = self._get_active_employees(team_id)
        weeks = self._get_iso_weeks(start_date, end_date)
        
        logger.info(f"Processing {len(employees)} employees across {len(weeks)} weeks")
        
        # Process each employee-week combination
        row = 2
        processed_count = 0
        skipped_count = 0
        
        for employee in employees:
            for week_start in weeks:
                compliance_data = self._evaluate_employee_week(employee, week_start)
                
                if compliance_data:  # Skip weeks with no activity
                    self._write_data_row(ws, row, compliance_data)
                    row += 1
                    processed_count += 1
                else:
                    skipped_count += 1
        
        # Apply formatting
        self._apply_formatting(ws, row - 1)
        
        # Generate filename and save
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = self.output_dir / f"JIRA_Compliance_Report_{timestamp}.xlsx"
        wb.save(output_path)
        
        logger.info(f"Compliance report saved: {output_path}")
        logger.info(f"Processed: {processed_count} records, Skipped: {skipped_count} (no activity)")
        
        return str(output_path)
    
    def _write_headers(self, ws: Worksheet):
        """Write column headers with formatting."""
        for col, header in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER
    
    def _write_data_row(self, ws: Worksheet, row: int, data: Dict[str, Any]):
        """Write single compliance data row."""
        # Column A: Employee Name
        cell = ws.cell(row=row, column=1, value=data['employee_name'])
        cell.font = self.DATA_FONT
        cell.alignment = Alignment(horizontal='left')
        cell.border = self.BORDER
        
        # Column B: Week Start Date
        cell = ws.cell(row=row, column=2, value=data['week_start'])
        cell.font = self.DATA_FONT
        cell.number_format = 'YYYY-MM-DD'
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.BORDER
        
        # Columns C-I: Compliance checks
        compliance_fields = [
            'status_hygiene',
            'cancellation',
            'update_frequency',
            'role_ownership',
            'documentation',
            'lifecycle',
            'zero_tolerance'
        ]
        
        for col, field in enumerate(compliance_fields, 3):
            cell = ws.cell(row=row, column=col, value=data[field])
            cell.font = self.DATA_FONT
            cell.alignment = Alignment(horizontal='left')
            cell.border = self.BORDER
        
        # Column J: Overall Compliance
        overall_cell = ws.cell(row=row, column=10, value=data['overall_compliance'])
        overall_cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        overall_cell.alignment = Alignment(horizontal='center')
        overall_cell.border = self.BORDER
        
        # Apply conditional formatting
        if data['overall_compliance'] == 'Pass':
            overall_cell.fill = self.PASS_FILL
        else:
            overall_cell.fill = self.FAIL_FILL
        
        # Column K: Auditor's Notes
        notes_cell = ws.cell(row=row, column=11, value=data['auditor_notes'])
        notes_cell.font = self.DATA_FONT
        notes_cell.alignment = Alignment(horizontal='left', wrap_text=True)
        notes_cell.border = self.BORDER
    
    def _evaluate_employee_week(
        self, 
        employee: JiraUser, 
        week_start: datetime
    ) -> Optional[Dict[str, Any]]:
        """
        Evaluate compliance for employee during specific week.
        
        Args:
            employee: JiraUser object
            week_start: Monday of the week
            
        Returns:
            Compliance data dictionary or None if no activity
        """
        week_end = week_start + timedelta(days=6)
        
        # Get employee's issues for the week
        issues = self._get_employee_issues(employee, week_start, week_end)
        
        if not issues:
            logger.debug(f"No activity for {employee.display_name} week of {week_start.date()}")
            return None
        
        logger.debug(f"Evaluating {employee.display_name} week of {week_start.date()} ({len(issues)} issues)")
        
        # Run all compliance checks
        results = {}
        for check_name, check in self.checks.items():
            try:
                # New checks return dict, legacy builder expects str
                check_result = check.evaluate(issues, employee)
                results[check_name] = self._map_result_to_legacy_format(check_result, check_name)
            except Exception as e:
                logger.error(f"Check {check_name} failed for {employee.display_name}: {e}")
                results[check_name] = "Error"
        
        # Calculate overall compliance
        overall = self._calculate_overall_compliance(results)
        
        # Generate auditor notes
        notes = self._generate_auditor_notes(results)
        
        return {
            'employee_name': employee.display_name,
            'week_start': week_start,
            'status_hygiene': results['status_hygiene'],
            'cancellation': results['cancellation'],
            'update_frequency': results['update_frequency'],
            'role_ownership': results['role_ownership'],
            'documentation': results['documentation'],
            'lifecycle': results['lifecycle'],
            'zero_tolerance': results['zero_tolerance'],
            'overall_compliance': overall,
            'auditor_notes': notes
        }

    def _map_result_to_legacy_format(self, result: Dict[str, Any], check_name: str) -> str:
        """Map new dictionary result to legacy string format."""
        status = result.get('status', 'NA')
        reason = result.get('reason', '')
        
        if status == 'Pass':
            if check_name == 'cancellation':
                return "No" # cancellations: No means Pass (no unauthorized cancellations)
            if check_name == 'zero_tolerance':
                return "No" # zero_tolerance: No means Pass (no violations)
            return "Yes"
            
        elif status == 'Fail':
            if check_name == 'cancellation':
                return f"Yes - {reason}" # cancellation: Yes means Fail (violation found)
            if check_name == 'zero_tolerance':
                return "Yes" # zero_tolerance: Yes means Fail
            return f"No - {reason}"
            
        return "NA"
        
        return {
            'employee_name': employee.display_name,
            'week_start': week_start,
            'status_hygiene': results['status_hygiene'],
            'cancellation': results['cancellation'],
            'update_frequency': results['update_frequency'],
            'role_ownership': results['role_ownership'],
            'documentation': results['documentation'],
            'lifecycle': results['lifecycle'],
            'zero_tolerance': results['zero_tolerance'],
            'overall_compliance': overall,
            'auditor_notes': notes
        }
    
    def _calculate_overall_compliance(self, results: Dict[str, str]) -> str:
        """
        Calculate overall pass/fail based on all checks.
        
        Logic: Fail if ANY check is "No" or contains "No -" or zero-tolerance violation
        """
        for check_name, result in results.items():
            # Check for "No" or "No - [reason]"
            if result.startswith("No"):
                return "Fail"
            
            # Zero-tolerance: "Yes" means violation found (inverted logic)
            if check_name == 'zero_tolerance' and result == "Yes":
                return "Fail"
            
            # Error in check
            if result == "Error":
                return "Fail"
        
        return "Pass"
    
    def _generate_auditor_notes(self, results: Dict[str, str]) -> str:
        """Generate human-readable notes from compliance results."""
        issues = []
        
        for check_name, result in results.items():
            # Extract reasons from "No - [reason]" format
            if result.startswith("No -"):
                issues.append(result[5:])  # Extract reason after "No - "
            elif result.startswith("Yes -") and check_name == 'cancellation':
                # Cancellation has "Yes - [details]" format
                issues.append(result[6:])
            elif result == "No" and check_name == 'cancellation':
                # No cancellation is good, skip
                pass
            elif result == "No":
                # Generic "No" without reason
                check_display = check_name.replace('_', ' ').title()
                issues.append(f"{check_display} failed")
            elif check_name == 'zero_tolerance' and result == "Yes":
                issues.append("Zero-tolerance violation detected")
            elif result == "Error":
                issues.append(f"{check_name.replace('_', ' ').title()} check error")
        
        return "; ".join(issues) if issues else "All checks passed"
    
    def _apply_formatting(self, ws: Worksheet, last_row: int):
        """Apply final formatting to worksheet."""
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Set column widths
        ws.column_dimensions['A'].width = 20  # Employee Name
        ws.column_dimensions['B'].width = 15  # Week Start Date
        ws.column_dimensions['C'].width = 22  # Status Hygiene
        ws.column_dimensions['D'].width = 28  # Cancellation
        ws.column_dimensions['E'].width = 20  # Updates
        ws.column_dimensions['F'].width = 28  # Roles
        ws.column_dimensions['G'].width = 32  # Documentation
        ws.column_dimensions['H'].width = 18  # Lifecycle
        ws.column_dimensions['I'].width = 22  # Zero-Tolerance
        ws.column_dimensions['J'].width = 22  # Overall Compliance
        ws.column_dimensions['K'].width = 50  # Auditor's Notes
    
    def _get_active_employees(self, team_id: Optional[int]) -> List[JiraUser]:
        """
        Get list of active employees for reporting period.
        
        Args:
            team_id: Optional team filter
            
        Returns:
            List of JiraUser objects
        """
        with get_session() as session:
            query_helpers = QueryHelpers(session)
            
            # Query active users from database
            query = session.query(JiraUser).filter(JiraUser.active == True)
            
            # TODO: Add team filtering when team-user mapping is available
            # if team_id:
            #     query = query.join(...).filter(Team.id == team_id)
            
            employees = query.order_by(JiraUser.display_name).all()
            
            logger.info(f"Found {len(employees)} active employees")
            return employees
    
    def _get_iso_weeks(self, start_date: datetime, end_date: datetime) -> List[datetime]:
        """
        Generate list of Monday dates (ISO week starts) in range.
        
        Args:
            start_date: Range start
            end_date: Range end
            
        Returns:
            List of datetime objects (all Mondays)
        """
        weeks = []
        
        # Go to the Monday of the week containing start_date
        current = start_date - timedelta(days=start_date.weekday())
        
        while current <= end_date:
            weeks.append(current)
            current += timedelta(days=7)
        
        logger.debug(f"Generated {len(weeks)} weeks from {weeks[0].date()} to {weeks[-1].date()}")
        return weeks
    
    def _get_employee_issues(
        self, 
        employee: JiraUser, 
        week_start: datetime, 
        week_end: datetime
    ) -> List[Dict[str, Any]]:
        """
        Fetch employee's JIRA issues for the week.
        
        Args:
            employee: JiraUser object
            week_start: Week start (Monday)
            week_end: Week end (Sunday)
            
        Returns:
            List of issue dictionaries with full data
        """
        try:
            # Build JQL query for employee's issues in the week
            # Issues where employee is assignee OR reporter, updated in the week
            jql = f'''
                (assignee = "{employee.account_id}" OR reporter = "{employee.account_id}")
                AND updated >= "{week_start.strftime('%Y-%m-%d')}"
                AND updated <= "{week_end.strftime('%Y-%m-%d')}"
            '''
            
            # Fetch issues with changelog and comments expanded
            issues = list(self.jira.fetch_issues(
                jql=jql.strip(),
                fields=['*all'],
                expand=['changelog', 'renderedFields'],
                max_results=100
            ))
            
            logger.debug(f"Fetched {len(issues)} issues for {employee.display_name}")
            return issues
            
        except Exception as e:
            logger.error(f"Failed to fetch issues for {employee.display_name}: {e}")
            return []


def generate_compliance_report(
    start_date: datetime,
    end_date: datetime,
    team_id: Optional[int] = None,
    output_path: Optional[str] = None
) -> str:
    """
    Convenience function to generate compliance report.
    
    Args:
        start_date: Report start date
        end_date: Report end date
        team_id: Optional team filter
        output_path: Optional output directory
        
    Returns:
        Path to generated report
    """
    from src.config_manager import ConfigManager
    
    config = ConfigManager()
    jira_client = JiraClient()
    
    builder = ComplianceReportBuilder(
        jira_client,
        output_dir=output_path or './outputs'
    )
    
    return builder.generate_report(start_date, end_date, team_id)
