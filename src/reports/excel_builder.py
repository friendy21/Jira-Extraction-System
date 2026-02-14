"""
Excel Dashboard Builder Module
Creates professional Excel dashboards with multiple sheets, charts, and formatting.
"""

from datetime import datetime 
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from src.config_manager import ConfigManager
from src.database.connection import get_session
from src.database.queries import QueryHelpers
from src.reports.charts import ChartBuilder
from src.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelBuilder:
    """
    Creates professional Excel dashboards with multiple sheets and charts.
    """
    
    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_ROW_FILL = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")
    
    BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    def __init__(self, output_path: str = None):
        """
        Initialize Excel builder.
        
        Args:
            output_path: Path for output file (optional)
        """
        self.config = ConfigManager()
        reports_config = self.config.get_reports_config()
        
        if output_path:
            self.output_path = Path(output_path)
        else:
            output_dir = Path(reports_config.get('output_dir', './outputs'))
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.output_path = output_dir / f"jira_dashboard_{timestamp}.xlsx"
        
        self.workbook = Workbook()
        self._setup_styles()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        logger.info(f"Excel builder initialized, output: {self.output_path}")
    
    def _setup_styles(self) -> None:
        """Setup named styles for the workbook."""
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = self.HEADER_FONT
        header_style.fill = self.HEADER_FILL
        header_style.alignment = Alignment(horizontal='center', vertical='center')
        
        # Number style
        number_style = NamedStyle(name="number")
        number_style.number_format = '#,##0'
        number_style.alignment = Alignment(horizontal='right')
        
        # Percent style
        percent_style = NamedStyle(name="percent")
        percent_style.number_format = '0.0%'
        percent_style.alignment = Alignment(horizontal='right')
        
        # Date style
        date_style = NamedStyle(name="date")
        date_style.number_format = 'YYYY-MM-DD'
        date_style.alignment = Alignment(horizontal='center')
        
        # Add styles to workbook if not present
        for style in [header_style, number_style, percent_style, date_style]:
            if style.name not in self.workbook.named_styles:
                self.workbook.add_named_style(style)
    
    def generate_dashboard(self, team_id: int = None) -> str:
        """
        Generate complete dashboard workbook.
        
        Args:
            team_id: Optional team filter
            
        Returns:
            Path to generated file
        """
        logger.info("Generating dashboard...")
        
        with get_session() as session:
            queries = QueryHelpers(session)
            
            # Create sheets
            self._create_executive_summary(queries, team_id)
            self._create_velocity_sheet(queries, team_id)
            self._create_sprint_analysis(queries, team_id)
            self._create_priority_sheet(queries, team_id)
            self._create_aging_sheet(queries, team_id)
            self._create_time_tracking_sheet(queries, team_id)
        
        # Save workbook
        self.workbook.save(self.output_path)
        logger.info(f"Dashboard saved to: {self.output_path}")
        
        return str(self.output_path)
    
    def _create_executive_summary(self, queries: QueryHelpers, team_id: int = None) -> None:
        """Create executive summary sheet."""
        ws = self.workbook.create_sheet("Executive Summary")
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Jira Dashboard - Executive Summary"
        title_cell.font = Font(bold=True, size=16, color="1F4E79")
        title_cell.alignment = Alignment(horizontal='center')
        
        # Generation timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=9, color="666666")
        
        row = 4
        
        # Key Metrics section
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E79")
        row += 2
        
        # Get metrics data
        if team_id:
            daily_metrics = queries.get_daily_metrics(team_id, days=30)
            time_tracking = queries.get_time_tracking_summary(team_id)
            velocity_data = queries.get_team_velocity(team_id, sprint_count=5)
        else:
            teams = queries.get_all_teams()
            daily_metrics = []
            velocity_data = []
            time_tracking = {'total_estimated_hours': 0, 'total_spent_hours': 0}
            
            for team in teams:
                daily_metrics.extend(queries.get_daily_metrics(team.id, days=30))
        
        # Calculate summary stats
        total_created = sum(m.get('tickets_created', 0) for m in daily_metrics)
        total_resolved = sum(m.get('tickets_resolved', 0) for m in daily_metrics)
        
        # Metrics table
        metrics = [
            ("Tickets Created (30 days)", total_created),
            ("Tickets Resolved (30 days)", total_resolved),
            ("Resolution Rate", f"{100 * total_resolved / max(total_created, 1):.1f}%"),
            ("Total Estimated Hours", time_tracking.get('total_estimated_hours', 0)),
            ("Total Logged Hours", time_tracking.get('total_spent_hours', 0)),
        ]
        
        for metric_name, metric_value in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = metric_value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        row += 2
        
        # Velocity section
        if velocity_data:
            ws[f'A{row}'] = "Recent Sprint Velocity"
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F4E79")
            row += 2
            
            # Headers
            headers = ['Sprint', 'Committed', 'Completed', 'Velocity', 'Completion %']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Data rows
            data_start_row = row
            for sprint in reversed(velocity_data):  # Show oldest first
                ws.cell(row=row, column=1, value=sprint.get('sprint_name', ''))
                ws.cell(row=row, column=2, value=sprint.get('points_committed', 0))
                ws.cell(row=row, column=3, value=sprint.get('points_completed', 0))
                ws.cell(row=row, column=4, value=sprint.get('velocity', 0))
                ws.cell(row=row, column=5, value=f"{sprint.get('completion_rate', 0)}%")
                row += 1
            
            # Add velocity chart
            if len(velocity_data) > 1:
                chart = ChartBuilder.create_velocity_chart(
                    ws,
                    start_row=data_start_row,
                    sprint_col=1,
                    committed_col=2,
                    completed_col=3,
                    num_sprints=len(velocity_data)
                )
                ws.add_chart(chart, f"G{data_start_row - 1}")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
    
    def _create_velocity_sheet(self, queries: QueryHelpers, team_id: int = None) -> None:
        """Create velocity analysis sheet."""
        ws = self.workbook.create_sheet("Velocity Analysis")
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "Team Velocity Analysis"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
        
        row = 3
        
        # Get all teams or specific team
        if team_id:
            teams = [queries.session.query(queries.get_all_teams.__self__).get(team_id)]
        else:
            teams = queries.get_all_teams()
        
        for team in teams:
            if not team:
                continue
            
            # Team header
            ws[f'A{row}'] = team.team_name
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Get velocity data
            velocity_data = queries.get_team_velocity(team.id, sprint_count=10)
            
            if velocity_data:
                # Headers
                headers = ['Sprint', 'Start Date', 'End Date', 'Committed', 'Completed', 'Velocity']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                row += 1
                
                # Data
                for sprint in reversed(velocity_data):
                    ws.cell(row=row, column=1, value=sprint.get('sprint_name', ''))
                    ws.cell(row=row, column=2, value=sprint.get('start_date').strftime('%Y-%m-%d') if sprint.get('start_date') else '')
                    ws.cell(row=row, column=3, value=sprint.get('end_date').strftime('%Y-%m-%d') if sprint.get('end_date') else '')
                    ws.cell(row=row, column=4, value=sprint.get('points_committed', 0))
                    ws.cell(row=row, column=5, value=sprint.get('points_completed', 0))
                    ws.cell(row=row, column=6, value=sprint.get('velocity', 0))
                    row += 1
                
                # Average velocity
                avg_velocity = sum(s.get('velocity', 0) for s in velocity_data) / len(velocity_data)
                ws[f'E{row}'] = "Average:"
                ws[f'E{row}'].font = Font(bold=True)
                ws[f'F{row}'] = round(avg_velocity, 1)
                ws[f'F{row}'].font = Font(bold=True)
            else:
                ws[f'A{row}'] = "No sprint data available"
                ws[f'A{row}'].font = Font(italic=True, color="666666")
            
            row += 2
        
        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_sprint_analysis(self, queries: QueryHelpers, team_id: int = None) -> None:
        """Create sprint analysis sheet."""
        ws = self.workbook.create_sheet("Sprint Analysis")
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "Sprint Analysis & Burndown"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
        
        row = 3
        
        # Get teams
        if team_id:
            teams = [t for t in queries.get_all_teams() if t.id == team_id]
        else:
            teams = queries.get_all_teams()
        
        for team in teams:
            # Team header
            ws[f'A{row}'] = team.team_name
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 2
            
            # Get recent sprints
            velocity_data = queries.get_team_velocity(team.id, sprint_count=5)
            
            for sprint in velocity_data[:3]:  # Show last 3 sprints in detail
                sprint_metrics = queries.get_sprint_metrics(sprint['sprint_id'])
                
                if sprint_metrics:
                    # Sprint name
                    ws[f'A{row}'] = sprint_metrics.get('sprint_name', 'Unknown Sprint')
                    ws[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    # Metrics grid
                    metrics = [
                        ('State', sprint_metrics.get('state', '')),
                        ('Total Issues', sprint_metrics.get('total_issues', 0)),
                        ('Completed Issues', sprint_metrics.get('completed_issues', 0)),
                        ('Total Points', sprint_metrics.get('total_points', 0)),
                        ('Completed Points', sprint_metrics.get('completed_points', 0)),
                        ('Completion %', f"{sprint_metrics.get('completion_percentage', 0)}%"),
                    ]
                    
                    for metric_name, metric_value in metrics:
                        ws[f'A{row}'] = metric_name
                        ws[f'B{row}'] = metric_value
                        row += 1
                    
                    if sprint_metrics.get('goal'):
                        ws[f'A{row}'] = "Goal"
                        ws[f'B{row}'] = sprint_metrics.get('goal', '')[:100]
                        row += 1
                    
                    row += 1
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def _create_priority_sheet(self, queries: QueryHelpers, team_id: int = None) -> None:
        """Create priority distribution sheet."""
        ws = self.workbook.create_sheet("Priority Analysis")
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "Priority Distribution"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
        
        row = 3
        
        # Get teams
        if team_id:
            teams = [t for t in queries.get_all_teams() if t.id == team_id]
        else:
            teams = queries.get_all_teams()
        
        for team in teams:
            # Team header
            ws[f'A{row}'] = team.team_name
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Get priority distribution
            priority_data = queries.get_priority_distribution(team.id)
            
            if priority_data:
                # Headers
                headers = ['Priority', 'Total', 'Open', 'Resolved']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                row += 1
                
                data_start_row = row
                
                # Data
                for item in priority_data:
                    ws.cell(row=row, column=1, value=item.get('priority', ''))
                    ws.cell(row=row, column=2, value=item.get('total_count', 0))
                    ws.cell(row=row, column=3, value=item.get('open_count', 0))
                    ws.cell(row=row, column=4, value=item.get('resolved_count', 0))
                    row += 1
                
                # Add pie chart
                if len(priority_data) > 1:
                    chart = ChartBuilder.create_pie_chart(
                        ws,
                        data_range=(3, data_start_row - 1, 3, row - 1),  # Open count
                        categories_range=(1, data_start_row, 1, row - 1),  # Priority names
                        title="Open Issues by Priority"
                    )
                    ws.add_chart(chart, f"F{data_start_row - 1}")
            else:
                ws[f'A{row}'] = "No priority data available"
                ws[f'A{row}'].font = Font(italic=True, color="666666")
            
            row += 2
        
        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_aging_sheet(self, queries: QueryHelpers, team_id: int = None) -> None:
        """Create ticket aging sheet."""
        ws = self.workbook.create_sheet("Ticket Aging")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Ticket Aging Analysis"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
        
        row = 3
        
        # Get teams
        if team_id:
            teams = [t for t in queries.get_all_teams() if t.id == team_id]
        else:
            teams = queries.get_all_teams()
        
        for team in teams:
            # Team header
            ws[f'A{row}'] = team.team_name
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Get aging data
            aging_data = queries.get_ticket_aging(team.id)
            
            if aging_data:
                # Headers
                headers = ['Age Bucket', 'Count']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                row += 1
                
                data_start_row = row
                
                # Data
                for item in aging_data:
                    ws.cell(row=row, column=1, value=item.get('bucket', ''))
                    ws.cell(row=row, column=2, value=item.get('count', 0))
                    row += 1
                
                # Add bar chart
                if len(aging_data) > 1:
                    chart = ChartBuilder.create_bar_chart(
                        ws,
                        data_range=(2, data_start_row - 1, 2, row - 1),
                        categories_range=(1, data_start_row, 1, row - 1),
                        title="Ticket Age Distribution"
                    )
                    ws.add_chart(chart, f"D{data_start_row - 1}")
            else:
                ws[f'A{row}'] = "No aging data available"
                ws[f'A{row}'].font = Font(italic=True, color="666666")
            
            row += 2
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
    
    def _create_time_tracking_sheet(self, queries: QueryHelpers, team_id: int = None) -> None:
        """Create time tracking sheet."""
        ws = self.workbook.create_sheet("Time Tracking")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Time Tracking Summary"
        ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
        
        row = 3
        
        # Get teams
        if team_id:
            teams = [t for t in queries.get_all_teams() if t.id == team_id]
        else:
            teams = queries.get_all_teams()
        
        for team in teams:
            # Team header
            ws[f'A{row}'] = team.team_name
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Get time tracking data
            time_data = queries.get_time_tracking_summary(team.id)
            
            if time_data:
                metrics = [
                    ('Total Estimated Hours', time_data.get('total_estimated_hours', 0)),
                    ('Total Logged Hours', time_data.get('total_spent_hours', 0)),
                    ('Variance (Hours)', time_data.get('variance_hours', 0)),
                    ('Accuracy %', f"{time_data.get('accuracy_percentage', 0):.1f}%" if time_data.get('accuracy_percentage') else 'N/A'),
                ]
                
                for metric_name, metric_value in metrics:
                    ws[f'A{row}'] = metric_name
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'] = metric_value
                    row += 1
            else:
                ws[f'A{row}'] = "No time tracking data available"
                ws[f'A{row}'].font = Font(italic=True, color="666666")
            
            row += 2
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def add_data_table(
        self,
        ws: Worksheet,
        data: List[Dict],
        start_row: int,
        columns: List[Tuple[str, str]],  # (header, key)
        table_name: str = None
    ) -> int:
        """
        Add a formatted data table to worksheet.
        
        Args:
            ws: Target worksheet
            data: List of dictionaries with data
            start_row: Starting row
            columns: List of (header, data_key) tuples
            table_name: Optional table name for Excel table
            
        Returns:
            Next available row after table
        """
        if not data:
            return start_row
        
        # Write headers
        for col, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col, (_, key) in enumerate(columns, 1):
                value = row_data.get(key, '')
                cell = ws.cell(row=row_idx, column=col, value=value)
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL
        
        end_row = start_row + len(data)
        end_col = len(columns)
        
        # Create Excel table if name provided
        if table_name:
            table_ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
            table = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        return end_row + 1


def generate_report(team_id: int = None, output_path: str = None) -> str:
    """
    Convenience function to generate a report.
    
    Args:
        team_id: Optional team filter
        output_path: Optional output file path
        
    Returns:
        Path to generated file
    """
    builder = ExcelBuilder(output_path)
    return builder.generate_dashboard(team_id)
